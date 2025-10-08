import json
import os
import csv
from models import db, PricingItem, AdminSetting
from utils.ai_helper import AIHelper
from utils.ai_pricing_extractor import AIPricingExtractor
from utils.pricing_service import PricingService
import requests

class PricingHelper:
    def __init__(self):
        self.ai_helper = AIHelper()
        self.ai_extractor = AIPricingExtractor()
        self.pricing_service = PricingService()
    
    def calculate_quote_pricing(self, quote):
        """Calculate pricing for a quote"""
        try:
            pricing_breakdown = {
                'materials': [],
                'labor': [],
                'total_materials': 0,
                'total_labor': 0,
                'total_cost': 0,
                'breakdown': {}
            }
            
            # Use AI's labour breakdown if available
            if quote.labour_breakdown:
                try:
                    labour_data = json.loads(quote.labour_breakdown)
                    if isinstance(labour_data, list):
                        total_project_hours = 0
                        total_project_cost = 0
                        day_rate = 0
                        
                        # First pass: calculate total hours and find day rate
                        for item in labour_data:
                            days = item.get('days', 0)
                            hours = item.get('hours', 0)
                            rate = item.get('day_rate', 0)
                            
                            if rate and not day_rate:
                                day_rate = float(rate)
                            
                            if days:
                                # Convert days to hours (assuming 8 hours per day)
                                total_project_hours += float(days) * 8
                            elif hours:
                                total_project_hours += float(hours)
                        
                        # Round total hours to nearest day for project cost
                        if total_project_hours > 0 and day_rate > 0:
                            total_project_days = max(0.5, round(total_project_hours / 8 * 2) / 2)  # Round to nearest 0.5 day
                            total_project_cost = total_project_days * day_rate
                            
                            # Add single labour item for the whole project
                            pricing_breakdown['labor'].append({
                                'item': 'Project Labour (All Tasks)',
                                'quantity': total_project_days,
                                'unit': 'days',
                                'unit_price': day_rate,
                                'total': total_project_cost,
                                'notes': f'Total project time: {total_project_hours:.1f} hours rounded to {total_project_days} days'
                            })
                            pricing_breakdown['total_labor'] = total_project_cost
                        else:
                            # Fallback: use individual task costs if total calculation fails
                            for item in labour_data:
                                cost = item.get('cost', 0)
                                days = item.get('days', 0)
                                day_rate = item.get('day_rate', 0)
                                
                                if cost:
                                    total_cost = float(cost)
                                elif days and day_rate:
                                    total_cost = float(days) * float(day_rate)
                                else:
                                    continue
                                
                                pricing_breakdown['labor'].append({
                                    'item': item.get('task', 'Labour'),
                                    'quantity': float(days) if days else 1,
                                    'unit': 'days',
                                    'unit_price': float(day_rate) if day_rate else total_cost,
                                    'total': total_cost
                                })
                                pricing_breakdown['total_labor'] += total_cost
                        
                        # Add material costs from AI products if available
                        if quote.recommended_products:
                            try:
                                products_data = json.loads(quote.recommended_products)
                                if isinstance(products_data, list):
                                    for product in products_data:
                                        # Use AI's pricing if available, otherwise try to get from database
                                        unit_price = product.get('unit_price', 0)
                                        total_price = product.get('total_price', 0)
                                        quantity = product.get('quantity', 1)
                                        
                                        # Safely convert quantity to float, handle text values
                                        try:
                                            if isinstance(quantity, str):
                                                # Try to extract number from string like "52.0 each" or "Allowance"
                                                import re
                                                quantity_match = re.search(r'(\d+\.?\d*)', str(quantity))
                                                if quantity_match:
                                                    quantity = float(quantity_match.group(1))
                                                else:
                                                    quantity = 1.0  # Default to 1 if no number found
                                            else:
                                                quantity = float(quantity)
                                        except (ValueError, TypeError):
                                            quantity = 1.0
                                        
                                        # Safely convert pricing to float
                                        try:
                                            unit_price = float(unit_price) if unit_price else 0
                                            total_price = float(total_price) if total_price else 0
                                        except (ValueError, TypeError):
                                            unit_price = 0
                                            total_price = 0
                                        
                                        if unit_price and total_price and unit_price > 0 and total_price > 0:
                                            # AI provided pricing - mark as estimated
                                            pricing_breakdown['materials'].append({
                                                'item': product.get('item', ''),
                                                'quantity': quantity,
                                                'unit': product.get('unit', 'each'),
                                                'unit_price': unit_price,
                                                'total': total_price,
                                                'part_number': product.get('part_number', ''),
                                                'pricing_source': 'ai_estimated',
                                                'is_estimated': True
                                            })
                                            pricing_breakdown['total_materials'] += total_price
                                        else:
                                            # Try to get real pricing from multiple sources
                                            product_name = product.get('item', '')
                                            unit_price = 0
                                            pricing_source = 'none'
                                            
                                            # Try multiple supplier names and product variations
                                            suppliers_to_try = self._get_suppliers_for_product(product_name)
                                            product_variations = self._get_product_name_variations(product_name)
                                            
                                            for supplier in suppliers_to_try:
                                                for variation in product_variations:
                                                    pricing_result = self.pricing_service.get_product_price(supplier, variation)
                                                    if pricing_result and pricing_result.get('success'):
                                                        unit_price = pricing_result.get('price', 0)
                                                        pricing_source = pricing_result.get('source', 'web_scraper')
                                                        break
                                                if unit_price > 0:
                                                    break
                                            
                                            # If still no pricing found, try database
                                            if unit_price == 0:
                                                pricing = self._get_item_price(None, None, product_name)
                                                if pricing:
                                                    unit_price = pricing
                                                    pricing_source = 'database'
                                            
                                            # If still no pricing, use realistic estimates
                                            if unit_price == 0:
                                                unit_price = self._get_realistic_estimate(product_name)
                                                pricing_source = 'estimated'
                                            
                                            total_cost = quantity * unit_price
                                            is_estimated = pricing_source in ['estimated', 'none']
                                            
                                            pricing_breakdown['materials'].append({
                                                'item': product.get('item', ''),
                                                'quantity': quantity,
                                                'unit': product.get('unit', 'each'),
                                                'unit_price': unit_price,
                                                'total': total_cost,
                                                'part_number': product.get('part_number', ''),
                                                'pricing_source': pricing_source,
                                                'is_estimated': is_estimated
                                            })
                                            pricing_breakdown['total_materials'] += total_cost
                            except Exception as e:
                                print(f"Error parsing AI products data: {e}")
                        
                        # Calculate total cost
                        pricing_breakdown['total_cost'] = pricing_breakdown['total_labor'] + pricing_breakdown['total_materials']
                        return pricing_breakdown
                except Exception as e:
                    print(f"Error parsing AI labour data: {e}")
                    # Fall back to old calculation if AI data is invalid
            
            # Calculate cabling costs
            if quote.cabling_type:
                cabling_cost = self._calculate_cabling_cost(quote)
                if cabling_cost:
                    pricing_breakdown['materials'].extend(cabling_cost['materials'])
                    pricing_breakdown['total_materials'] += cabling_cost['total']
                    pricing_breakdown['breakdown']['cabling'] = cabling_cost
            
            # Calculate WiFi costs
            if quote.wifi_requirements:
                wifi_cost = self._calculate_wifi_cost(quote)
                if wifi_cost:
                    pricing_breakdown['materials'].extend(wifi_cost['materials'])
                    pricing_breakdown['total_materials'] += wifi_cost['total']
                    pricing_breakdown['breakdown']['wifi'] = wifi_cost
            
            # Calculate CCTV costs
            if quote.cctv_requirements:
                cctv_cost = self._calculate_cctv_cost(quote)
                if cctv_cost:
                    pricing_breakdown['materials'].extend(cctv_cost['materials'])
                    pricing_breakdown['total_materials'] += cctv_cost['total']
                    pricing_breakdown['breakdown']['cctv'] = cctv_cost
            
            # Calculate door entry costs
            if quote.door_entry_requirements:
                door_cost = self._calculate_door_entry_cost(quote)
                if door_cost:
                    pricing_breakdown['materials'].extend(door_cost['materials'])
                    pricing_breakdown['total_materials'] += door_cost['total']
                    pricing_breakdown['breakdown']['door_entry'] = door_cost
            
            # Calculate labor costs
            labor_cost = self._calculate_labor_cost(quote)
            pricing_breakdown['labor'] = labor_cost['labor']
            pricing_breakdown['total_labor'] = labor_cost['total']
            pricing_breakdown['breakdown']['labor'] = labor_cost
            
            # Calculate total
            pricing_breakdown['total_cost'] = pricing_breakdown['total_materials'] + pricing_breakdown['total_labor']
            
            return pricing_breakdown
            
        except Exception as e:
            print(f"Error calculating pricing: {e}")
            return None
    
    def _get_suppliers_for_product(self, product_name):
        """Get likely suppliers for a product based on product name"""
        product_lower = product_name.lower()
        
        suppliers = []
        
        # UniFi products
        if any(keyword in product_lower for keyword in ['unifi', 'ubiquiti', 'dream machine', 'u6', 'u7', 'g5', 'g6', 'protect']):
            suppliers.extend(['Ubiquiti UniFi', 'Ubiquiti', 'Ubiquiti Networks'])
        
        # CCTV products
        if any(keyword in product_lower for keyword in ['camera', 'cctv', 'protect', 'nvr']):
            suppliers.extend(['Ubiquiti UniFi Protect', 'Ubiquiti UniFi', 'Ubiquiti'])
        
        # Network equipment
        if any(keyword in product_lower for keyword in ['switch', 'router', 'firewall', 'poe']):
            suppliers.extend(['Ubiquiti UniFi', 'Ubiquiti', 'Cisco', 'Netgear'])
        
        # Cabling products
        if any(keyword in product_lower for keyword in ['cat6', 'cat6a', 'cable', 'patch panel', 'keystone', 'faceplate']):
            suppliers.extend(['Connectix', 'Panduit', 'Commscope', 'Belden'])
        
        # Fiber products
        if any(keyword in product_lower for keyword in ['om4', 'fiber', 'sfp', 'multimode']):
            suppliers.extend(['Connectix', 'Panduit', 'Commscope'])
        
        # Default suppliers if no specific match
        if not suppliers:
            suppliers = ['Ubiquiti UniFi', 'Connectix', 'Generic']
        
        return suppliers
    
    def _get_product_name_variations(self, product_name):
        """Get variations of product name for better matching"""
        variations = [product_name]  # Always include original
        
        # Remove common descriptive text
        cleaned = product_name.lower()
        
        # Remove parenthetical descriptions
        import re
        cleaned = re.sub(r'\([^)]*\)', '', cleaned).strip()
        if cleaned != product_name.lower():
            variations.append(cleaned)
        
        # Extract key product identifiers
        if 'dream machine' in cleaned:
            variations.extend(['Dream Machine', 'UDM', 'UDM-Pro', 'Dream Machine Pro'])
        elif 'u7-pro' in cleaned:
            variations.extend(['U7-Pro', 'U7 Pro', 'WiFi 7 AP'])
        elif 'u6' in cleaned:
            variations.extend(['U6-Pro', 'U6-Lite', 'U6-LR'])
        elif 'g5' in cleaned or 'g6' in cleaned:
            variations.extend(['G5-Bullet', 'G5-Dome', 'G6-Bullet', 'G6-Dome'])
        elif 'switch' in cleaned and 'poe' in cleaned:
            variations.extend(['PoE Switch', 'Switch PoE', 'Managed Switch'])
        elif 'cat6' in cleaned:
            variations.extend(['Cat6 Cable', 'CAT6', 'Cat6A'])
        elif 'patch panel' in cleaned:
            variations.extend(['Patch Panel', 'CAT6 Patch Panel', '24-port Patch Panel'])
        
        return variations
    
    def _get_realistic_estimate(self, product_name):
        """Get realistic pricing estimates for common products"""
        product_lower = product_name.lower()
        
        # UniFi products
        if 'dream machine' in product_lower:
            return 279.0 if 'pro' in product_lower else 199.0
        elif 'u7-pro' in product_lower:
            return 167.0
        elif 'u6-pro' in product_lower:
            return 125.0
        elif 'u6-lite' in product_lower:
            return 89.0
        elif 'g5-bullet' in product_lower:
            return 179.0
        elif 'g5-dome' in product_lower:
            return 199.0
        elif 'g6' in product_lower:
            return 149.0
        elif 'nvr' in product_lower:
            return 399.0
        
        # Network equipment
        elif 'switch' in product_lower and 'poe' in product_lower:
            if '48' in product_lower:
                return 899.0
            elif '24' in product_lower:
                return 399.0
            else:
                return 299.0
        elif 'switch' in product_lower:
            if '48' in product_lower:
                return 299.0
            elif '24' in product_lower:
                return 199.0
            else:
                return 149.0
        
        # Cabling products
        elif 'cat6' in product_lower and 'cable' in product_lower:
            return 45.0  # Per 305m reel
        elif 'patch panel' in product_lower:
            if '48' in product_lower:
                return 35.0
            elif '24' in product_lower:
                return 25.0
            else:
                return 20.0
        elif 'keystone' in product_lower:
            return 3.0
        elif 'faceplate' in product_lower:
            return 2.0
        elif 'patch lead' in product_lower or 'patch cable' in product_lower:
            return 5.0
        
        # Fiber products
        elif 'om4' in product_lower and 'cable' in product_lower:
            return 8.0  # Per meter
        elif 'sfp' in product_lower:
            return 25.0
        
        # General equipment
        elif 'pdu' in product_lower:
            return 89.0
        elif 'rack' in product_lower:
            return 199.0
        elif 'consumables' in product_lower:
            return 50.0  # Lump sum estimate
        
        # Default estimate
        return 25.0
    
    def _calculate_cabling_cost(self, quote):
        """Calculate cabling material costs"""
        try:
            # Estimate cable length based on building size and floors
            estimated_length = self._estimate_cable_length(quote)
            
            # Get pricing for cabling materials
            cable_price = self._get_item_price('cabling', 'cable', quote.cabling_type)
            if not cable_price:
                cable_price = 0.5  # Default price per meter
            
            # Calculate materials needed
            materials = [
                {
                    'item': f'{quote.cabling_type.upper()} Cable',
                    'quantity': estimated_length,
                    'unit': 'meters',
                    'unit_price': cable_price,
                    'total': estimated_length * cable_price
                },
                {
                    'item': 'Network Points',
                    'quantity': quote.number_of_rooms * 2,  # 2 points per room average
                    'unit': 'points',
                    'unit_price': 25,  # Cost per point including faceplate, module, etc.
                    'total': quote.number_of_rooms * 2 * 25
                },
                {
                    'item': 'Patch Panel',
                    'quantity': max(1, quote.number_of_rooms // 24),  # 24 ports per panel
                    'unit': 'units',
                    'unit_price': 150,
                    'total': max(1, quote.number_of_rooms // 24) * 150
                },
                {
                    'item': 'Network Switch',
                    'quantity': max(1, quote.number_of_rooms // 48),  # 48 ports per switch
                    'unit': 'units',
                    'unit_price': 300,
                    'total': max(1, quote.number_of_rooms // 48) * 300
                }
            ]
            
            total_cost = sum(item['total'] for item in materials)
            
            return {
                'materials': materials,
                'total': total_cost
            }
            
        except Exception as e:
            print(f"Error calculating cabling cost: {e}")
            return None
    
    def _calculate_wifi_cost(self, quote):
        """Calculate WiFi equipment costs"""
        try:
            # Estimate number of access points needed
            ap_count = self._estimate_access_points(quote)
            
            materials = [
                {
                    'item': 'UniFi WiFi 6 Access Point',
                    'quantity': ap_count,
                    'unit': 'units',
                    'unit_price': 150,
                    'total': ap_count * 150
                },
                {
                    'item': 'PoE Injector',
                    'quantity': ap_count,
                    'unit': 'units',
                    'unit_price': 25,
                    'total': ap_count * 25
                }
            ]
            
            # Add UniFi controller if multiple APs
            if ap_count > 2:
                materials.append({
                    'item': 'UniFi Dream Machine',
                    'quantity': 1,
                    'unit': 'units',
                    'unit_price': 300,
                    'total': 300
                })
            
            total_cost = sum(item['total'] for item in materials)
            
            return {
                'materials': materials,
                'total': total_cost
            }
            
        except Exception as e:
            print(f"Error calculating WiFi cost: {e}")
            return None
    
    def _calculate_cctv_cost(self, quote):
        """Calculate CCTV equipment costs"""
        try:
            # Estimate number of cameras needed
            camera_count = self._estimate_camera_count(quote)
            
            materials = [
                {
                    'item': 'UniFi Camera',
                    'quantity': camera_count,
                    'unit': 'units',
                    'unit_price': 200,
                    'total': camera_count * 200
                },
                {
                    'item': 'NVR Storage',
                    'quantity': 1,
                    'unit': 'units',
                    'unit_price': 400,
                    'total': 400
                }
            ]
            
            total_cost = sum(item['total'] for item in materials)
            
            return {
                'materials': materials,
                'total': total_cost
            }
            
        except Exception as e:
            print(f"Error calculating CCTV cost: {e}")
            return None
    
    def _calculate_door_entry_cost(self, quote):
        """Calculate door entry system costs"""
        try:
            materials = [
                {
                    'item': 'Paxton Net2 Entry System',
                    'quantity': 1,
                    'unit': 'units',
                    'unit_price': 800,
                    'total': 800
                },
                {
                    'item': 'Door Reader',
                    'quantity': 2,  # Entry and exit
                    'unit': 'units',
                    'unit_price': 150,
                    'total': 300
                },
                {
                    'item': 'Door Controller',
                    'quantity': 1,
                    'unit': 'units',
                    'unit_price': 200,
                    'total': 200
                }
            ]
            
            total_cost = sum(item['total'] for item in materials)
            
            return {
                'materials': materials,
                'total': total_cost
            }
            
        except Exception as e:
            print(f"Error calculating door entry cost: {e}")
            return None
    
    def _calculate_labor_cost(self, quote):
        """Calculate labor costs"""
        try:
            labor_items = []

            cabling_hours = 0
            cabinet_hours = 0
            testing_hours = 0

            outlet_count = max(quote.number_of_rooms or 1, 1) * 2
            multiplier_map = {
                'cat5e': 0.35,
                'cat6': 0.45,
                'cat6a': 0.6,
                'fiber': 0.8
            }

            if quote.project_description:
                desc_lower = quote.project_description.lower()
                if 'double' in desc_lower and 'cat' in desc_lower:
                    outlet_count = max(outlet_count, 20)
            if quote.cabling_type:
                per_outlet_hours = multiplier_map.get(quote.cabling_type.lower(), 0.4)
            else:
                per_outlet_hours = 0.4
            cabling_hours += outlet_count * per_outlet_hours

            building_multiplier = 1.0
            if quote.building_size:
                if quote.building_size > 500:
                    building_multiplier = 1.2
                elif quote.building_size > 1000:
                    building_multiplier = 1.4
            cabling_hours *= building_multiplier

            cabinet_hours = 2 if quote.project_description and 'cab' in quote.project_description.lower() else 0

            testing_hours = max(outlet_count * 0.1, 1)

            wifi_hours = 0.0
            if quote.wifi_requirements:
                wifi_hours = max(quote.number_of_rooms or 1, 1) * 0.5 + 1
            cctv_hours = 0.0
            if quote.cctv_requirements:
                cctv_hours = max((quote.number_of_rooms or 1), 1) * 0.75 + 1
            door_hours = 0.0
            if quote.door_entry_requirements:
                door_hours = 3

            total_hours = cabling_hours + cabinet_hours + testing_hours + wifi_hours + cctv_hours + door_hours
            total_hours = max(total_hours, 4)

            try:
                day_rate_setting = AdminSetting.query.filter_by(key='day_rate').first()
            except Exception:
                day_rate_setting = None
            day_rate_value = None
            if day_rate_setting:
                try:
                    day_rate_value = float(day_rate_setting.value)
                except (TypeError, ValueError):
                    day_rate_value = None

            hourly_rate = 90
            if day_rate_value and day_rate_value > 0:
                hourly_rate = day_rate_value / 8.0

            labor_items.append({
                'item': 'Structured Cabling Labour',
                'quantity': round(total_hours, 1),
                'unit': 'hours',
                'days': round(total_hours / 8.0, 2),
                'unit_price': hourly_rate,
                'total': round(total_hours * hourly_rate, 2)
            })

            if cabinet_hours:
                labor_items.append({
                    'item': 'Cabinet Assembly & Commissioning',
                    'quantity': round(cabinet_hours, 1),
                    'unit': 'hours',
                    'days': round(cabinet_hours / 8.0, 2),
                    'unit_price': hourly_rate,
                    'total': round(cabinet_hours * hourly_rate, 2)
                })

            if testing_hours:
                labor_items.append({
                    'item': 'Testing & Certification',
                    'quantity': round(testing_hours, 1),
                    'unit': 'hours',
                    'days': round(testing_hours / 8.0, 2),
                    'unit_price': hourly_rate,
                    'total': round(testing_hours * hourly_rate, 2)
                })

            if wifi_hours:
                labor_items.append({
                    'item': 'WiFi Deployment',
                    'quantity': round(wifi_hours, 1),
                    'unit': 'hours',
                    'days': round(wifi_hours / 8.0, 2),
                    'unit_price': hourly_rate,
                    'total': round(wifi_hours * hourly_rate, 2)
                })

            if cctv_hours:
                labor_items.append({
                    'item': 'CCTV Installation',
                    'quantity': round(cctv_hours, 1),
                    'unit': 'hours',
                    'days': round(cctv_hours / 8.0, 2),
                    'unit_price': hourly_rate,
                    'total': round(cctv_hours * hourly_rate, 2)
                })

            if door_hours:
                labor_items.append({
                    'item': 'Door Entry Installation',
                    'quantity': round(door_hours, 1),
                    'unit': 'hours',
                    'days': round(door_hours / 8.0, 2),
                    'unit_price': hourly_rate,
                    'total': round(door_hours * hourly_rate, 2)
                })

            total_cost = sum(item['total'] for item in labor_items)

            return {
                'labor': labor_items,
                'total': total_cost
            }

        except Exception as e:
            print(f"Error calculating labor cost: {e}")
            return None
    
    def _estimate_cable_length(self, quote):
        """Estimate cable length needed"""
        if quote.building_size:
            # Rough estimate: 2 meters of cable per square meter of building
            return int(quote.building_size * 2)
        else:
            # Fallback: estimate based on rooms and floors
            return quote.number_of_rooms * quote.number_of_floors * 50
    
    def _estimate_access_points(self, quote):
        """Estimate number of WiFi access points needed"""
        if quote.building_size:
            # One AP per 100 square meters
            return max(1, int(quote.building_size / 100))
        else:
            # Fallback: one per floor, minimum 2
            return max(2, quote.number_of_floors)
    
    def _estimate_camera_count(self, quote):
        """Estimate number of CCTV cameras needed"""
        if quote.building_size:
            # One camera per 50 square meters
            return max(2, int(quote.building_size / 50))
        else:
            # Fallback: 2 cameras minimum, one per floor
            return max(2, quote.number_of_floors)
    
    def _estimate_installation_time(self, quote):
        """Estimate total installation time in hours"""
        base_hours = 8
        
        if quote.cabling_type:
            base_hours += 4
        if quote.wifi_requirements:
            base_hours += 2
        if quote.cctv_requirements:
            base_hours += 3
        if quote.door_entry_requirements:
            base_hours += 2
        
        # Scale by building size
        if quote.building_size:
            size_factor = min(quote.building_size / 200, 2)
            base_hours = int(base_hours * (1 + size_factor))
        
        return base_hours
    
    def _get_item_price(self, category, subcategory, product_name):
        """Get price for a specific item from database"""
        try:
            item = PricingItem.query.filter_by(
                category=category,
                subcategory=subcategory,
                product_name=product_name
            ).first()
            
            if item:
                return item.cost_per_unit
            
            return None
            
        except Exception as e:
            print(f"Error getting item price: {e}")
            return None
    
    def get_product_pricing(self, products):
        """Get pricing for multiple products"""
        try:
            pricing_data = {}
            
            for product in products:
                category = product.get('category')
                name = product.get('name')
                
                # Try to get from database first
                db_price = self._get_item_price(category, None, name)
                
                if db_price:
                    pricing_data[name] = {
                        'source': 'database',
                        'price': db_price
                    }
                else:
                    # Try to get from web
                    web_price = self.ai_helper.get_web_pricing(name, category)
                    if web_price:
                        pricing_data[name] = {
                            'source': 'web',
                            'price': web_price
                        }
            
            return pricing_data
            
        except Exception as e:
            print(f"Error getting product pricing: {e}")
            return {}
    
    def import_excel_pricing(self, file_path):
        """Import pricing from Excel or CSV file using AI extraction"""
        try:
            # First try AI-powered extraction for any format
            ai_extracted_data = self.ai_extractor.extract_pricing_from_spreadsheet(file_path)
            
            if ai_extracted_data:
                return self._save_ai_extracted_pricing(ai_extracted_data, file_path)
            
            # Fallback to traditional method if AI extraction fails
            return self._traditional_import(file_path)
            
        except Exception as e:
            print(f"Error importing pricing: {e}")
            return 0
    
    def _save_ai_extracted_pricing(self, pricing_data, file_path):
        """Save AI-extracted pricing data to database"""
        try:
            imported_count = 0
            filename = os.path.basename(file_path)
            
            for item in pricing_data:
                try:
                    pricing_item = PricingItem(
                        category=item.get('category', 'other'),
                        subcategory=item.get('subcategory'),
                        product_name=item.get('product_name'),
                        description=item.get('description'),
                        unit=item.get('unit', 'piece'),
                        cost_per_unit=item.get('cost_per_unit', 0),
                        supplier=item.get('supplier'),
                        part_number=item.get('part_number'),
                        source=f'ai_extracted_{filename}'
                    )
                    
                    db.session.add(pricing_item)
                    imported_count += 1
                    
                except Exception as e:
                    print(f"Error saving AI extracted item: {e}")
                    continue
            
            db.session.commit()
            return imported_count
            
        except Exception as e:
            print(f"Error saving AI extracted pricing: {e}")
            return 0
    
    def _traditional_import(self, file_path):
        """Traditional spreadsheet import method"""
        try:
            imported_count = 0
            
            # Try to read as CSV first (for .csv files)
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    for row in reader:
                        try:
                            pricing_item = PricingItem(
                                category=str(row.get('Category', '')),
                                subcategory=str(row.get('Subcategory', '')),
                                product_name=str(row.get('Product Name', '')),
                                description=str(row.get('Description', '')),
                                unit=str(row.get('Unit', 'piece')),
                                cost_per_unit=float(row.get('Cost Per Unit', 0)),
                                supplier=str(row.get('Supplier', '')),
                                part_number=str(row.get('Part Number', '')),
                                source='csv'
                            )
                            
                            db.session.add(pricing_item)
                            imported_count += 1
                            
                        except Exception as e:
                            print(f"Error importing row: {e}")
                            continue
            else:
                # For Excel files, we'll need openpyxl
                try:
                    from openpyxl import load_workbook
                    workbook = load_workbook(file_path)
                    worksheet = workbook.active
                    
                    # Get headers from first row
                    headers = [cell.value for cell in worksheet[1]]
                    
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        try:
                            # Create dictionary from row data
                            row_data = dict(zip(headers, row))
                            
                            pricing_item = PricingItem(
                                category=str(row_data.get('Category', '')),
                                subcategory=str(row_data.get('Subcategory', '')),
                                product_name=str(row_data.get('Product Name', '')),
                                description=str(row_data.get('Description', '')),
                                unit=str(row_data.get('Unit', 'piece')),
                                cost_per_unit=float(row_data.get('Cost Per Unit', 0)),
                                supplier=str(row_data.get('Supplier', '')),
                                part_number=str(row_data.get('Part Number', '')),
                                source='excel'
                            )
                            
                            db.session.add(pricing_item)
                            imported_count += 1
                            
                        except Exception as e:
                            print(f"Error importing row: {e}")
                            continue
                            
                except ImportError:
                    print("openpyxl not available for Excel import")
                    return 0
            
            db.session.commit()
            return imported_count
            
        except Exception as e:
            print(f"Error in traditional import: {e}")
            return 0
    
    def update_pricing_from_web(self):
        """Update pricing from web sources"""
        try:
            # Get items that need updating
            items = PricingItem.query.filter(
                PricingItem.source != 'web'
            ).limit(50).all()  # Update in batches
            
            updated_count = 0
            
            for item in items:
                try:
                    # Get updated price from web
                    web_price = self.ai_helper.get_web_pricing(
                        item.product_name, 
                        item.category
                    )
                    
                    if web_price and web_price != item.cost_per_unit:
                        item.cost_per_unit = web_price
                        item.source = 'web'
                        updated_count += 1
                        
                except Exception as e:
                    print(f"Error updating price for {item.product_name}: {e}")
                    continue
            
            db.session.commit()
            return updated_count
            
        except Exception as e:
            print(f"Error updating pricing from web: {e}")
            return 0
    
    def find_missing_pricing(self, quote_requirements):
        """Find products that need pricing and suggest web searches"""
        try:
            missing_products = []
            
            # Extract product names from quote requirements
            if quote_requirements.get('wifi_requirements'):
                missing_products.extend(self._extract_products_from_text(quote_requirements['wifi_requirements']))
            
            if quote_requirements.get('cctv_requirements'):
                missing_products.extend(self._extract_products_from_text(quote_requirements['cctv_requirements']))
            
            if quote_requirements.get('door_entry_requirements'):
                missing_products.extend(self._extract_products_from_text(quote_requirements['door_entry_requirements']))
            
            # Check which products are missing from database
            missing_from_db = []
            for product in missing_products:
                existing = PricingItem.query.filter(
                    PricingItem.product_name.ilike(f"%{product}%")
                ).first()
                
                if not existing:
                    missing_from_db.append(product)
            
            # Use AI to analyze missing products and suggest web searches
            if missing_from_db:
                analysis = self.ai_extractor.analyze_pricing_gaps(missing_from_db)
                return analysis
            
            return []
            
        except Exception as e:
            print(f"Error finding missing pricing: {e}")
            return []
    
    def _extract_products_from_text(self, text):
        """Extract product names from text using AI"""
        try:
            if not text or not self.ai_extractor.openai_client:
                return []
            
            prompt = f"""
            Extract product names, models, and specifications from this text:
            "{text}"
            
            Focus on:
            - Cable types (Cat5e, Cat6, Cat6a, Fiber, etc.)
            - Equipment models (Unifi, Net2, specific camera models, etc.)
            - Quantities and specifications
            
            Return as a simple list of product names, one per line.
            """
            
            response = self.ai_extractor.openai_client.chat.completions.create(
                model="gpt-5",
                messages=[
                    {"role": "system", "content": "Extract product names from text. Return only a simple list."},
                    {"role": "user", "content": prompt}
                ],
                # temperature=0.3,  # GPT-5 only supports default temperature
                max_completion_tokens=500
            )
            
            result_text = response.choices[0].message.content
            products = [line.strip() for line in result_text.split('\n') if line.strip()]
            
            return products
            
        except Exception as e:
            print(f"Error extracting products from text: {e}")
            return []
    
    def get_web_pricing_for_product(self, product_name, category=None):
        """Get current web pricing for a specific product"""
        try:
            return self.ai_extractor.get_web_pricing(product_name, category)
        except Exception as e:
            print(f"Error getting web pricing for {product_name}: {e}")
            return None
    
    def import_from_spreadsheet_folder(self, folder_path="pricing_spreadsheets"):
        """Import pricing from all spreadsheets in the designated folder"""
        try:
            if not os.path.exists(folder_path):
                print(f"Folder {folder_path} does not exist")
                return 0
            
            total_imported = 0
            
            # Get all spreadsheet files
            spreadsheet_extensions = ['.csv', '.xlsx', '.xls', '.ods']
            files = []
            
            for file in os.listdir(folder_path):
                if any(file.lower().endswith(ext) for ext in spreadsheet_extensions):
                    files.append(os.path.join(folder_path, file))
            
            # Import each file
            for file_path in files:
                try:
                    print(f"Importing pricing from: {file_path}")
                    imported_count = self.import_excel_pricing(file_path)
                    total_imported += imported_count
                    print(f"Imported {imported_count} items from {os.path.basename(file_path)}")
                    
                except Exception as e:
                    print(f"Error importing {file_path}: {e}")
                    continue
            
            return total_imported
            
        except Exception as e:
            print(f"Error importing from spreadsheet folder: {e}")
            return 0
